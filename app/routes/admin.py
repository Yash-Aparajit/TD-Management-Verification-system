"""
Admin: TD management (lines, FG codes, TD items), export. No developer-only tools.
"""
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, abort
from flask_login import current_user
from ..extensions import db
from ..models import Line, FGCode, TDItem, AuditLog
from ..decorators import admin_required
from ..services.audit_service import log_td_create, log_td_update, log_td_deactivate
from ..utils.validators import normalize_fg_code, normalize_unit, normalize_whitespace
from ..config import ITEMS_PER_PAGE, TD_ITEMS_PER_PAGE
import io
from datetime import datetime

admin_bp = Blueprint("admin", __name__)


@admin_bp.route("/")
@admin_required
def dashboard():
    from ..models import Line, FGCode, TDItem, Verification
    from datetime import datetime, timedelta
    # Statistics
    total_lines = Line.query.count()
    active_lines = Line.query.filter_by(is_active=True).count()
    total_fg_codes = FGCode.query.count()
    active_fg_codes = FGCode.query.filter_by(is_active=True).count()
    total_td_items = TDItem.query.count()
    active_td_items = TDItem.query.filter_by(is_active=True).count()
    total_verifications = Verification.query.count()
    recent_verifications = Verification.query.order_by(Verification.verified_at.desc()).limit(5).all()
    today_verifications = Verification.query.filter(
        Verification.verified_at >= datetime.utcnow().replace(hour=0, minute=0, second=0)
    ).count()
    return render_template(
        "admin/dashboard.html",
        total_lines=total_lines,
        active_lines=active_lines,
        total_fg_codes=total_fg_codes,
        active_fg_codes=active_fg_codes,
        total_td_items=total_td_items,
        active_td_items=active_td_items,
        total_verifications=total_verifications,
        today_verifications=today_verifications,
        recent_verifications=recent_verifications,
    )


# ---- Lines ----
@admin_bp.route("/lines")
@admin_required
def lines_list():
    page = request.args.get("page", 1, type=int)
    q = Line.query.order_by(Line.code)
    if request.args.get("active_only"):
        q = q.filter(Line.is_active == True)
    pagination = q.paginate(page=page, per_page=ITEMS_PER_PAGE)
    return render_template("admin/lines_list.html", pagination=pagination)


@admin_bp.route("/lines/create", methods=["GET", "POST"])
@admin_required
def line_create():
    if request.method == "GET":
        return render_template("admin/line_form.html", line=None)
    code = normalize_whitespace(request.form.get("code"))
    name = normalize_whitespace(request.form.get("name"))
    if not code:
        flash("Line code is required.", "danger")
        return render_template("admin/line_form.html", line=None)
    if Line.query.filter_by(code=code).first():
        flash("Line with this code already exists.", "danger")
        return render_template("admin/line_form.html", line=None)
    line = Line(code=code, name=name or code, updated_by_id=current_user.id)
    db.session.add(line)
    db.session.commit()
    log_td_create(current_user.id, current_user.username, "line", line.id)
    flash("Line created.", "success")
    return redirect(url_for("admin.lines_list"))


@admin_bp.route("/lines/<int:line_id>/edit", methods=["GET", "POST"])
@admin_required
def line_edit(line_id):
    line = Line.query.get_or_404(line_id)
    if request.method == "GET":
        return render_template("admin/line_form.html", line=line)
    code = normalize_whitespace(request.form.get("code"))
    name = normalize_whitespace(request.form.get("name"))
    if not code:
        flash("Line code is required.", "danger")
        return render_template("admin/line_form.html", line=line)
    other = Line.query.filter(Line.code == code, Line.id != line_id).first()
    if other:
        flash("Another line with this code exists.", "danger")
        return render_template("admin/line_form.html", line=line)
    line.code = code
    line.name = name or code
    line.updated_by_id = current_user.id
    db.session.commit()
    log_td_update(current_user.id, current_user.username, "line", line.id)
    flash("Line updated.", "success")
    return redirect(url_for("admin.lines_list"))


@admin_bp.route("/lines/<int:line_id>/deactivate", methods=["POST"])
@admin_required
def line_deactivate(line_id):
    line = Line.query.get_or_404(line_id)
    line.is_active = False
    line.updated_by_id = current_user.id
    db.session.commit()
    log_td_deactivate(current_user.id, current_user.username, "line", line.id)
    flash("Line deactivated.", "success")
    return redirect(url_for("admin.lines_list"))


@admin_bp.route("/lines/<int:line_id>/activate", methods=["POST"])
@admin_required
def line_activate(line_id):
    line = Line.query.get_or_404(line_id)
    line.is_active = True
    line.updated_by_id = current_user.id
    db.session.commit()
    log_td_update(current_user.id, current_user.username, "line", line.id, details="activated")
    flash("Line activated.", "success")
    return redirect(url_for("admin.lines_list"))


# ---- FG Codes ----
@admin_bp.route("/fg")
@admin_required
def fg_list():
    page = request.args.get("page", 1, type=int)
    line_id = request.args.get("line_id", type=int)
    search = (request.args.get("q") or "").strip().upper()
    q = FGCode.query.join(Line).order_by(Line.code, FGCode.code)
    if line_id:
        q = q.filter(FGCode.line_id == line_id)
    if search:
        q = q.filter(FGCode.code.ilike(f"%{search}%"))
    if request.args.get("active_only"):
        q = q.filter(FGCode.is_active == True)
    pagination = q.paginate(page=page, per_page=ITEMS_PER_PAGE)
    lines = Line.query.filter_by(is_active=True).order_by(Line.code).all()
    return render_template("admin/fg_list.html", pagination=pagination, lines=lines)


@admin_bp.route("/fg/create", methods=["GET", "POST"])
@admin_required
def fg_create():
    lines = Line.query.filter_by(is_active=True).order_by(Line.code).all()
    if not lines:
        flash("Create at least one line first.", "warning")
        return redirect(url_for("admin.lines_list"))
    if request.method == "GET":
        return render_template("admin/fg_form.html", fg=None, lines=lines)
    line_id = request.form.get("line_id", type=int)
    code = normalize_fg_code(request.form.get("code"))
    name = normalize_whitespace(request.form.get("name"))
    if not line_id or not code:
        flash("Line and FG code are required.", "danger")
        return render_template("admin/fg_form.html", fg=None, lines=lines)
    if FGCode.query.filter_by(line_id=line_id, code=code).first():
        flash("FG code already exists for this line.", "danger")
        return render_template("admin/fg_form.html", fg=None, lines=lines)
    fg = FGCode(line_id=line_id, code=code, name=name or code, updated_by_id=current_user.id)
    db.session.add(fg)
    db.session.commit()
    log_td_create(current_user.id, current_user.username, "fg_code", fg.id)
    flash("FG code created.", "success")
    return redirect(url_for("admin.fg_list"))


@admin_bp.route("/fg/<int:fg_id>/edit", methods=["GET", "POST"])
@admin_required
def fg_edit(fg_id):
    fg = FGCode.query.get_or_404(fg_id)
    lines = Line.query.filter_by(is_active=True).order_by(Line.code).all()
    if request.method == "GET":
        return render_template("admin/fg_form.html", fg=fg, lines=lines)
    code = normalize_fg_code(request.form.get("code"))
    name = normalize_whitespace(request.form.get("name"))
    if not code:
        flash("FG code is required.", "danger")
        return render_template("admin/fg_form.html", fg=fg, lines=lines)
    other = FGCode.query.filter(FGCode.line_id == fg.line_id, FGCode.code == code, FGCode.id != fg_id).first()
    if other:
        flash("Another FG code with this code exists for this line.", "danger")
        return render_template("admin/fg_form.html", fg=fg, lines=lines)
    fg.code = code
    fg.name = name or code
    fg.updated_by_id = current_user.id
    db.session.commit()
    log_td_update(current_user.id, current_user.username, "fg_code", fg.id)
    flash("FG code updated.", "success")
    return redirect(url_for("admin.fg_list"))


@admin_bp.route("/fg/<int:fg_id>/deactivate", methods=["POST"])
@admin_required
def fg_deactivate(fg_id):
    fg = FGCode.query.get_or_404(fg_id)
    fg.is_active = False
    fg.updated_by_id = current_user.id
    db.session.commit()
    log_td_deactivate(current_user.id, current_user.username, "fg_code", fg.id)
    flash("FG code deactivated.", "success")
    return redirect(url_for("admin.fg_list"))


@admin_bp.route("/fg/<int:fg_id>/activate", methods=["POST"])
@admin_required
def fg_activate(fg_id):
    fg = FGCode.query.get_or_404(fg_id)
    fg.is_active = True
    fg.updated_by_id = current_user.id
    db.session.commit()
    log_td_update(current_user.id, current_user.username, "fg_code", fg.id, details="activated")
    flash("FG code activated.", "success")
    return redirect(url_for("admin.fg_list"))


# ---- TD Items (per FG) ----
@admin_bp.route("/fg/<int:fg_id>/td")
@admin_required
def td_list(fg_id):
    fg = FGCode.query.get_or_404(fg_id)
    page = request.args.get("page", 1, type=int)
    search = (request.args.get("q") or "").strip()
    item_type = request.args.get("type")
    q = TDItem.query.filter_by(fg_id=fg_id).order_by(TDItem.item_code)
    if search:
        q = q.filter(
            db.or_(
                TDItem.item_code.ilike(f"%{search}%"),
                TDItem.item_name.ilike(f"%{search}%"),
            )
        )
    if item_type and item_type in ("child_part", "consumable"):
        q = q.filter(TDItem.item_type == item_type)
    if request.args.get("active_only"):
        q = q.filter(TDItem.is_active == True)
    pagination = q.paginate(page=page, per_page=TD_ITEMS_PER_PAGE)
    return render_template("admin/td_list.html", fg=fg, pagination=pagination)


@admin_bp.route("/fg/<int:fg_id>/td/create", methods=["GET", "POST"])
@admin_required
def td_create(fg_id):
    fg = FGCode.query.get_or_404(fg_id)
    if request.method == "GET":
        return render_template("admin/td_item_form.html", fg=fg, item=None)
    item_code = normalize_whitespace(request.form.get("item_code"))
    item_name = normalize_whitespace(request.form.get("item_name"))
    item_type = request.form.get("item_type") or "child_part"
    if item_type not in ("child_part", "consumable"):
        item_type = "child_part"
    try:
        quantity = float(request.form.get("quantity") or 0)
    except ValueError:
        quantity = 0
    unit = normalize_unit(request.form.get("unit"))
    if not item_code:
        flash("Item code is required.", "danger")
        return render_template("admin/td_item_form.html", fg=fg, item=None)
    if quantity < 0:
        flash("Quantity must be >= 0.", "danger")
        return render_template("admin/td_item_form.html", fg=fg, item=None)
    if TDItem.query.filter_by(fg_id=fg_id, item_code=item_code).first():
        flash("Item code already exists for this FG.", "danger")
        return render_template("admin/td_item_form.html", fg=fg, item=None)
    item = TDItem(
        fg_id=fg_id,
        item_code=item_code,
        item_name=item_name or item_code,
        item_type=item_type,
        quantity=quantity,
        unit=unit,
        updated_by_id=current_user.id,
    )
    db.session.add(item)
    db.session.commit()
    log_td_create(current_user.id, current_user.username, "td_item", item.id)
    flash("TD item created.", "success")
    return redirect(url_for("admin.td_list", fg_id=fg_id))


@admin_bp.route("/fg/<int:fg_id>/td/<int:item_id>/edit", methods=["GET", "POST"])
@admin_required
def td_edit(fg_id, item_id):
    fg = FGCode.query.get_or_404(fg_id)
    item = TDItem.query.filter_by(id=item_id, fg_id=fg_id).first_or_404()
    if request.method == "GET":
        return render_template("admin/td_item_form.html", fg=fg, item=item)
    item_code = normalize_whitespace(request.form.get("item_code"))
    item_name = normalize_whitespace(request.form.get("item_name"))
    item_type = request.form.get("item_type") or "child_part"
    if item_type not in ("child_part", "consumable"):
        item_type = "child_part"
    try:
        quantity = float(request.form.get("quantity") or 0)
    except ValueError:
        quantity = 0
    unit = normalize_unit(request.form.get("unit"))
    if not item_code:
        flash("Item code is required.", "danger")
        return render_template("admin/td_item_form.html", fg=fg, item=item)
    if quantity < 0:
        flash("Quantity must be >= 0.", "danger")
        return render_template("admin/td_item_form.html", fg=fg, item=item)
    other = TDItem.query.filter(TDItem.fg_id == fg_id, TDItem.item_code == item_code, TDItem.id != item_id).first()
    if other:
        flash("Another item with this code exists for this FG.", "danger")
        return render_template("admin/td_item_form.html", fg=fg, item=item)
    # Edit conflict: if form's updated_at is older than current DB value, someone else saved in between
    submitted_at = request.form.get("updated_at")
    if submitted_at and item.updated_at:
        try:
            from datetime import datetime as dt
            raw = submitted_at.replace("Z", "").strip()[:19]
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S"):
                try:
                    submitted_dt = dt.strptime(raw, fmt)
                    db_val = item.updated_at.replace(tzinfo=None) if getattr(item.updated_at, "tzinfo", None) else item.updated_at
                    if submitted_dt < db_val:
                        flash("This record was modified by someone else. Please refresh and try again.", "warning")
                        return render_template("admin/td_item_form.html", fg=fg, item=item)
                    break
                except ValueError:
                    continue
        except Exception:
            pass
    item.item_code = item_code
    item.item_name = item_name or item_code
    item.item_type = item_type
    item.quantity = quantity
    item.unit = unit
    item.updated_by_id = current_user.id
    db.session.commit()
    log_td_update(current_user.id, current_user.username, "td_item", item.id)
    flash("TD item updated.", "success")
    return redirect(url_for("admin.td_list", fg_id=fg_id))


@admin_bp.route("/fg/<int:fg_id>/td/<int:item_id>/deactivate", methods=["POST"])
@admin_required
def td_deactivate(fg_id, item_id):
    item = TDItem.query.filter_by(id=item_id, fg_id=fg_id).first_or_404()
    item.is_active = False
    item.updated_by_id = current_user.id
    db.session.commit()
    log_td_deactivate(current_user.id, current_user.username, "td_item", item.id)
    flash("TD item deactivated.", "success")
    return redirect(url_for("admin.td_list", fg_id=fg_id))


@admin_bp.route("/fg/<int:fg_id>/td/<int:item_id>/activate", methods=["POST"])
@admin_required
def td_activate(fg_id, item_id):
    item = TDItem.query.filter_by(id=item_id, fg_id=fg_id).first_or_404()
    item.is_active = True
    item.updated_by_id = current_user.id
    db.session.commit()
    log_td_update(current_user.id, current_user.username, "td_item", item.id, details="activated")
    flash("TD item activated.", "success")
    return redirect(url_for("admin.td_list", fg_id=fg_id))


# ---- Export ----
@admin_bp.route("/export/td/<int:fg_id>")
@admin_required
def export_td(fg_id):
    fg = FGCode.query.get_or_404(fg_id)
    items = TDItem.query.filter_by(fg_id=fg_id).order_by(TDItem.item_code).all()
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        flash("Excel export not available. Install openpyxl.", "danger")
        return redirect(url_for("admin.td_list", fg_id=fg_id))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TD Items"
    ws.append(["Item Code", "Item Name", "Type", "Quantity", "Unit", "Updated At", "Updated By", "Active"])
    for row in ws.iter_rows(min_row=1, max_row=1):
        for c in row:
            c.font = Font(bold=True)
    for it in items:
        updated_by = it.updated_by.username if it.updated_by else ""
        ws.append([
            it.item_code,
            it.item_name,
            it.item_type,
            float(it.quantity),
            it.unit,
            it.updated_at.strftime("%Y-%m-%d %H:%M") if it.updated_at else "",
            updated_by,
            "Yes" if it.is_active else "No",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"TD_{fg.code}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
    )


@admin_bp.route("/export/audit-logs")
@admin_required
def export_audit_logs():
    from_date = request.args.get("from")
    to_date = request.args.get("to")
    q = AuditLog.query.order_by(AuditLog.created_at.desc())
    if from_date:
        try:
            q = q.filter(AuditLog.created_at >= datetime.strptime(from_date, "%Y-%m-%d"))
        except ValueError:
            pass
    if to_date:
        try:
            from datetime import timedelta
            end = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
            q = q.filter(AuditLog.created_at < end)
        except ValueError:
            pass
    logs = q.limit(10000).all()
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        flash("Excel export not available. Install openpyxl.", "danger")
        return redirect(url_for("admin.dashboard"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Logs"
    ws.append(["Time", "User", "Action", "Resource", "Resource ID", "Details", "IP"])
    for row in ws.iter_rows(min_row=1, max_row=1):
        for c in row:
            c.font = Font(bold=True)
    for log in reversed(logs):
        ws.append([
            log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "",
            log.username or "",
            log.action,
            log.resource or "",
            log.resource_id or "",
            log.details or "",
            log.ip_address or "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"audit_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx",
    )


@admin_bp.route("/audit-logs")
@admin_required
def audit_logs():
    page = request.args.get("page", 1, type=int)
    q = AuditLog.query.order_by(AuditLog.created_at.desc())
    pagination = q.paginate(page=page, per_page=ITEMS_PER_PAGE)
    return render_template("admin/audit_logs.html", pagination=pagination)


@admin_bp.route("/export/verifications")
@admin_required
def export_verifications():
    from ..models import Verification, VerificationItem
    from_date = request.args.get("from")
    to_date = request.args.get("to")
    q = Verification.query.order_by(Verification.verified_at.desc())
    if from_date:
        try:
            q = q.filter(Verification.verified_at >= datetime.strptime(from_date, "%Y-%m-%d"))
        except ValueError:
            pass
    if to_date:
        try:
            from datetime import timedelta
            end = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
            q = q.filter(Verification.verified_at < end)
        except ValueError:
            pass
    verifications = q.limit(5000).all()
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        flash("Excel export not available. Install openpyxl.", "danger")
        return redirect(url_for("admin.dashboard"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verifications"
    ws.append(["Verified At", "FG Code", "Operator", "Notes", "Item Code", "Expected", "Actual", "Unit"])
    for row in ws.iter_rows(min_row=1, max_row=1):
        for c in row:
            c.font = Font(bold=True)
    for ver in verifications:
        fg_code = ver.fg_code.code if ver.fg_code else ""
        op_name = ver.operator.username if ver.operator else ""
        for vi in ver.items:
            item_code = vi.td_item.item_code if vi.td_item else ""
            ws.append([
                ver.verified_at.strftime("%Y-%m-%d %H:%M") if ver.verified_at else "",
                fg_code,
                op_name,
                (ver.notes or "")[:200],
                item_code,
                float(vi.expected_quantity),
                float(vi.actual_quantity),
                vi.unit,
            ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"verifications_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx",
    )
